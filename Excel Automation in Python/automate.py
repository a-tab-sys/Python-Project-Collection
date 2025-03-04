# Program: reduce the price in the transactions.xlsx file by 10% with automation and create a chart.

# openpyxl is a module
# BarChart, Reference are classes
import openpyxl as xl     #required for excel sheet processing
from openpyxl.chart import BarChart, Reference  #required for creating the chart
import datetime

def process_workbook(filename):  #placing code in a function so it can be reused on any number of files

    print(xl.__version__)

    #load workbook and returns workbook object
    wb = xl.load_workbook(filename)   
    
    #even though your script and transactions.xlsx are in the same folder, Python may be running from a different directory.
    #file_path=r"C:\Users\Hasaan\Documents\Python-Project-Collection\Automation with Python\transactions.xlsx"
    #wb = xl.load_workbook(file_path)

    #access a specific sheet. sheet is an object
    sheet = wb['Sheet1']

    #access specific cells with coordinates.
    cell = sheet['a1']  #another way...
    samecell=sheet.cell(1,1)
    # print(cell.value)
    # print(samecell.value)

    #now we need to iterate to the third column value in each row to get the price. we will take the price and multiply it by 0.9
    #see how many rows are in the spreadsheet
    # print(sheet.max_row)
    maxrow=sheet.max_row

    #create for loop to iterate tough the rows of the excel sheet. go from 2-4 because row 1 in excel is the headings
    for row in range(2, maxrow+1):
        #print(row)
        cell=sheet.cell(row, 3) #go to each row in the third column
        #print(cell.value)
        corrected_price=cell.value*0.9
        corrected_price_cell=sheet.cell(row,4)   #create new column to place the correct prices. this returns a cell object.
        corrected_price_cell.value=corrected_price  #update the cells with the correct prices

    #creating the chart
    #use reference classs to select values for our chart. selecting column 4 cells, from row 2 to 4
    #values will be an object that holds all or 4th column values
    values=Reference(sheet, 
            min_row=2, 
            max_row=maxrow,
            min_col=4,
            max_col=4)

    #create an instance of bar chart class and store in an object. 
    chart=BarChart()
    chart.add_data(values)  #pass our values
    sheet.add_chart(chart, 'e2')  #add chart to our sheet with coordinate of where it should be placed (top left of chart at e2)

    # save the workbook as a new file instead of overwriting the original by appending a timestamp or a suffix to the filename   
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    new_filename = f"corrected_transactions_{timestamp}.xlsx"
    wb.save(new_filename)

    #to overwrite same file
    # wb.save(filename)


process_workbook('transactions.xlsx')

